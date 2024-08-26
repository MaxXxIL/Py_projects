import csv
from openpyxl import load_workbook
import openpyxl
from collections import defaultdict


def load_sheet_to_dict(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = defaultdict(list)

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
        label, image = row
        data[label].append(image)

    return data


def find_unique_images(dict1, dict2, sheet1_name, sheet2_name):
    unique_images = defaultdict(lambda: defaultdict(list))
    all_unique_images = []

    for label in set(dict1.keys()) | set(dict2.keys()):
        images1 = set(dict1.get(label, []))
        images2 = set(dict2.get(label, []))

        unique_to_1 = images1 - images2
        unique_to_2 = images2 - images1

        if unique_to_1:
            unique_images[label]['list1'] = list(unique_to_1)
            all_unique_images.extend([(label, img, sheet1_name) for img in unique_to_1])
        if unique_to_2:
            unique_images[label]['list2'] = list(unique_to_2)
            all_unique_images.extend([(label, img, sheet2_name) for img in unique_to_2])

    return unique_images, all_unique_images


def save_to_csv(data, filename):
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Label', 'Image', 'Original Sheet'])  # Header
        writer.writerows(data)


# Example usage:
file_path = 'D:\\Clean Data evaluation\\CleanLab\\KYEC Test\\Cleanlab step1.xlsx'
sheet_names = ['Org labels', 'Operator labels']  # Replace with your actual sheet names
sheet1_name = 'Org labels'  # Replace with your actual first sheet name
sheet2_name = 'Operator labels'
output_path = 'D:\\Clean Data evaluation\\\CleanLab\\\KYEC Test\\differences.csv'

# Load data from Excel sheets
dict1 = load_sheet_to_dict(file_path, sheet1_name)
dict2 = load_sheet_to_dict(file_path, sheet2_name)
# Compare lists and save differences to CSV
unique_images, all_unique_images = find_unique_images(dict1, dict2, sheet1_name, sheet2_name)
save_to_csv(all_unique_images, output_path)
print(f"\nResults have been saved to {output_path}")
x=1