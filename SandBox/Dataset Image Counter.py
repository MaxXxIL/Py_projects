import os
import csv
from collections import defaultdict
from tkinter import messagebox, filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def count_images_in_datasets(root_folder):
    dataset_counts = defaultdict(lambda: defaultdict(int))
    dataset_set = set()

    for dataset in os.listdir(root_folder):
        dataset_path = os.path.join(root_folder, dataset)
        if os.path.isdir(dataset_path):
            dataset_set.add(dataset)
            for label in os.listdir(dataset_path):
                label_folder = os.path.join(dataset_path, label)
                if os.path.isdir(label_folder):
                    image_count = sum(1 for file in os.listdir(label_folder) if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')))
                    dataset_counts[label][dataset] = image_count

    return dataset_counts, sorted(dataset_set)

import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

def count_images_in_datasets(root_folder):
    dataset_counts = defaultdict(lambda: defaultdict(int))
    dataset_set = set()

    for dataset in os.listdir(root_folder):
        dataset_path = os.path.join(root_folder, dataset)
        if os.path.isdir(dataset_path):
            dataset_set.add(dataset)
            for label in os.listdir(dataset_path):
                label_folder = os.path.join(dataset_path, label)
                if os.path.isdir(label_folder):
                    image_count = sum(1 for file in os.listdir(label_folder) if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')))
                    dataset_counts[label][dataset] = image_count

    return dataset_counts, sorted(dataset_set)


def write_excel(dataset_counts, datasets, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Counts"

    # Write header
    header = ['Label'] + datasets + ['Total']
    ws.append(header)

    # Style for header
    header_fill = PatternFill(start_color="FF1A1A1A", end_color="FF1A1A1A", fill_type="solid")  # Almost black
    header_font = Font(color="FFFFFF", bold=True)  # White and bold
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Calculate dataset totals
    dataset_totals = defaultdict(int)
    label_totals = defaultdict(int)

    # Sort labels alphabetically
    sorted_labels = sorted(dataset_counts.keys())

    # Write data rows
    for label in sorted_labels:
        counts = dataset_counts[label]
        row = [label]
        label_total = 0
        for dataset in datasets:
            count = counts.get(dataset, 0)
            row.append(count)
            label_total += count
            dataset_totals[dataset] += count
        row.append(label_total)  # Add label total
        label_totals[label] = label_total
        ws.append(row)

        # Color-code the total cell
        total_cell = ws.cell(row=ws.max_row, column=len(datasets) + 2)
        if label_total < 20:
            total_cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # Red
        elif label_total < 200:
            total_cell.fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Yellow

    # Write total row
    total_row = ['Total']
    grand_total = 0
    for dataset in datasets:
        total = dataset_totals[dataset]
        total_row.append(total)
        grand_total += total
    total_row.append(grand_total)
    ws.append(total_row)

    # Make total row bold
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Add borders to all cells
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(datasets) + 2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)


def main():
    root_folder = filedialog.askdirectory()
    output_file = root_folder + "\\parreto.xlsx"

    dataset_counts, datasets = count_images_in_datasets(root_folder)
    write_excel(dataset_counts, datasets, output_file)
    print(f"CSV file '{output_file}' has been created successfully.")


if __name__ == "__main__":
    main()