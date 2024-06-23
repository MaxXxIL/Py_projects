import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook

def count_files_in_folder(folder_path):
    count = 0
    for _, _, files in os.walk(folder_path):
        count += len(files)
    return count

def save_to_excel(folder_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Folder Contents"

    ws['A1'] = 'Folder Name'
    ws['B1'] = 'Number of Files'

    row = 2
    for root, _, _ in os.walk(folder_path):
        folder_name = os.path.basename(root)
        file_count = count_files_in_folder(root)
        ws.cell(row=row, column=1, value=folder_name)
        ws.cell(row=row, column=2, value=file_count)
        row += 1

    excel_filename = os.path.join(folder_path, 'folder_contents.xlsx')
    wb.save(excel_filename)
    print(f"Excel file saved as '{excel_filename}'")

def select_folder():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    folder_path = filedialog.askdirectory(title="Select a folder")
    if folder_path:
        save_to_excel(folder_path)
    else:
        print("No folder selected.")

if __name__ == "__main__":
    select_folder()